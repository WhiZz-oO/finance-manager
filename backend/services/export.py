"""
Excel export service using openpyxl.
Generates a multi-sheet workbook with financial data.
"""

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models.transaction import Transaction
from models.account import Account
from models.category import Category
from services.analytics import (
    get_account_balance, get_summary, get_monthly_analytics, get_category_breakdown
)

# Color palette
HEADER_BG = "1E1B4B"   # deep indigo
INCOME_BG = "D1FAE5"   # emerald light
EXPENSE_BG = "FEE2E2"  # rose light
TRANSFER_BG = "E0E7FF" # indigo light
WHITE = "FFFFFF"
DARK = "111827"
BORDER_COLOR = "E5E7EB"


def _header_style(cell, bg=HEADER_BG):
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def _auto_width(ws, min_width=12):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=min_width)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 4, min_width)


def generate_excel(db: Session, month: int = None, year: int = None) -> bytes:
    wb = Workbook()
    today = date.today()
    month = month or today.month
    year = year or today.year
    start = date(year, month, 1)
    end = date(year, month + 1, 1) if month < 12 else date(year, 12, 31)

    # ─── Sheet 1: Summary ────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells("A1:C1")
    ws_sum["A1"] = f"📊 Finance Summary — {start.strftime('%B %Y')}"
    ws_sum["A1"].font = Font(bold=True, size=14, color=HEADER_BG)
    ws_sum["A1"].alignment = Alignment(horizontal="center")

    accounts = db.query(Account).filter(Account.is_active == True).all()
    row = 3
    ws_sum.cell(row, 1, "Account").font = Font(bold=True)
    ws_sum.cell(row, 2, "Type").font = Font(bold=True)
    ws_sum.cell(row, 3, "Balance (₹)").font = Font(bold=True)
    row += 1
    for acc in accounts:
        bal = get_account_balance(acc.id, db)
        ws_sum.cell(row, 1, acc.name)
        ws_sum.cell(row, 2, acc.account_type)
        ws_sum.cell(row, 3, bal).number_format = '₹#,##0.00'
        row += 1

    row += 1
    summary = get_summary(db, start, end)
    for label, val in [
        ("Total Income", summary["total_income"]),
        ("Total Expenses", summary["total_expense"]),
        ("Net Cash Flow", summary["net"]),
        ("Savings Rate", f"{summary['savings_rate']}%"),
    ]:
        ws_sum.cell(row, 1, label).font = Font(bold=True)
        ws_sum.cell(row, 2, val)
        row += 1

    _auto_width(ws_sum)

    # ─── Sheet 2: Transactions ────────────────────────────────────────────
    ws_txn = wb.create_sheet("Transactions")
    headers = ["Ref", "Date", "Account", "Type", "Category", "Merchant", "Description", "Amount (₹)"]
    for col, h in enumerate(headers, 1):
        cell = ws_txn.cell(1, col, h)
        _header_style(cell)

    txns = (
        db.query(Transaction)
        .filter(
            Transaction.is_deleted == False,
            Transaction.transaction_date >= start,
            Transaction.transaction_date < end,
        )
        .order_by(Transaction.transaction_date.desc())
        .all()
    )
    for row_i, txn in enumerate(txns, 2):
        account = db.query(Account).filter(Account.id == txn.account_id).first()
        cat = db.query(Category).filter(Category.id == txn.category_id).first() if txn.category_id else None
        sign = "+" if txn.transaction_type in ("income", "refund") else "-"
        bg = INCOME_BG if sign == "+" else EXPENSE_BG
        fill = PatternFill("solid", fgColor=bg)

        data = [
            txn.reference or "",
            str(txn.transaction_date),
            account.name if account else "",
            txn.transaction_type.upper(),
            cat.name if cat else "",
            txn.merchant or "",
            txn.description or "",
            txn.amount if sign == "+" else -txn.amount,
        ]
        for col_i, val in enumerate(data, 1):
            cell = ws_txn.cell(row_i, col_i, val)
            cell.fill = fill
            cell.border = _thin_border()
            if col_i == 8:
                cell.number_format = '₹#,##0.00'

    _auto_width(ws_txn)

    # ─── Sheet 3: Monthly Analytics ──────────────────────────────────────
    ws_monthly = wb.create_sheet("Monthly Analytics")
    headers = ["Month", "Income (₹)", "Expenses (₹)", "Net (₹)", "Savings Rate"]
    for col, h in enumerate(headers, 1):
        cell = ws_monthly.cell(1, col, h)
        _header_style(cell)

    monthly = get_monthly_analytics(db, year)
    for row_i, m in enumerate(monthly, 2):
        ws_monthly.cell(row_i, 1, m["month_name"])
        ws_monthly.cell(row_i, 2, m["total_income"]).number_format = '₹#,##0.00'
        ws_monthly.cell(row_i, 3, m["total_expense"]).number_format = '₹#,##0.00'
        net_cell = ws_monthly.cell(row_i, 4, m["net"])
        net_cell.number_format = '₹#,##0.00'
        if m["net"] >= 0:
            net_cell.font = Font(color="059669")
        else:
            net_cell.font = Font(color="DC2626")
        ws_monthly.cell(row_i, 5, f"{m['savings_rate']}%")

    _auto_width(ws_monthly)

    # ─── Sheet 4: Category Breakdown ─────────────────────────────────────
    ws_cat = wb.create_sheet("Category Breakdown")
    headers = ["Category", "Amount (₹)", "% of Total"]
    for col, h in enumerate(headers, 1):
        cell = ws_cat.cell(1, col, h)
        _header_style(cell)

    cats = get_category_breakdown(db, start, end)
    for row_i, c in enumerate(cats, 2):
        ws_cat.cell(row_i, 1, f"{c['icon']} {c['category']}")
        ws_cat.cell(row_i, 2, c["amount"]).number_format = '₹#,##0.00'
        ws_cat.cell(row_i, 3, f"{c['percentage']}%")

    _auto_width(ws_cat)

    # ─── Serialize ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
